import openpyxl, os
from openpyxl.utils import get_column_letter
from tabulate import tabulate

workbook = openpyxl.load_workbook('mess_expense.xlsx') if 'mess_expense.xlsx' in os.listdir() else openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Expense'

if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
    headers = ["ID", "Month", "Room Rent", "Electric Bill", "Food", "Total"]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    workbook.save('mess_expense.xlsx')

def update_ids():
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1, value=row-1)
    workbook.save('mess_expense.xlsx')

def is_table_empty():
    return worksheet.max_row == 1

def get_integer_input(prompt):
    while True:
        try:
            value = int(input(prompt))
            return value
        except ValueError:
            print("Please enter a valid integer")

def get_value_from_table(column, data_id):
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == data_id:
            return worksheet.cell(row=row, column=column).value
        print(f"No data found for id {data_id}")
        return None

def view_data():
    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    headers = ["ID", "Month", "Room Rent", "Electric Bill", "Food", "Total"]
    table = tabulate(rows, headers, tablefmt='pretty')
    print("\n" + table + "\n")

def add_data():
    while True:
        month = input("Enter month: ")
        room_rent = get_integer_input("Enter room rent: ")
        electic_units = get_integer_input("Enter eletric units used: ")
        food = get_integer_input("Enter food cost: ")
        if (electic_units == 0): electic_bill = 0
        else: electic_bill = electic_units*8.5 + 25
        total = room_rent + electic_bill + food
        new_row = [worksheet.max_row, month, room_rent, electic_bill, food, total]
        worksheet.append(new_row)
        workbook.save('mess_expense.xlsx')
        print("Entry added!")
        view_data()
        repeat = input("Do you want to add another entry (y/n): ")
        if repeat.lower() != 'y':
            break

def update_data():
    while True: 
        view_data()
        data_id = int(input("Enter data id to be updated: "))
        print("1. Month")
        print("2. Room Rent")
        print("3. Electric Bill")
        print("4. Food")
        update_choice = input("Select update field: ")
        row_to_update = None
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == data_id:
                row_to_update = row
                break
        if row_to_update is None: 
            print("Invalid ID!")
            continue

        old_room_rent = worksheet.cell(row=row_to_update, column=3).value
        old_eletric_bill = worksheet.cell(row=row_to_update, column=4).value
        old_food = worksheet.cell(row=row_to_update, column=5).value

        match update_choice:
            case '1':
                new_month = input("Enter month: ")
                worksheet.cell(row=row_to_update, column=2, value=new_month)
            case '2':
                new_room_rent = get_integer_input("Enter room rent: ")
                new_total = new_room_rent + old_eletric_bill + old_food
                worksheet.cell(row=row_to_update, column=3, value=new_room_rent)
                worksheet.cell(row=row_to_update, column=6, value=new_total)
            case '3':
                new_electric_units = get_integer_input("Enter electric units used: ")
                if (new_electric_units != 0):  new_electric_bill = new_electric_units * 8.5 + 25
                else: new_electric_bill = 0
                new_total = old_room_rent + new_electric_bill + old_food
                worksheet.cell(row=row_to_update, column=4, value=new_electric_bill)
                worksheet.cell(row=row_to_update, column=6, value=new_total)
            case '4':
                new_food = get_integer_input("Enter food cost: ")
                new_total = old_room_rent + old_eletric_bill + new_food
                worksheet.cell(row=row_to_update, column=5, value=new_food)
                worksheet.cell(row=row_to_update, column=6, value=new_total)
            case _:
                print("Invalid input!")
        workbook.save('mess_expense.xlsx')
        print("Entry updated!")
        repeat = input("Do you want to update another entry(y/n): ")
        if repeat.lower() != 'y':
            break



def delete_data():
    while not is_table_empty():
        view_data()
        data_id = int(input("Enter data id to be deleted: "))
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == data_id:
                worksheet.delete_rows(row)
                break
        update_ids()
        workbook.save('mess_expence.xlsx')
        print("Data deleted!")
        if is_table_empty():
            print("No more data remaining")
        else: 
            repeat = input("Do you want to delete another entry(y/n): ")
            if repeat.lower() != 'y':
                break

def main():
    print("|| Welcome to Mess Expence Tracker ||")
    while True:
        a = 1
        print(f"{a}. Add Data")
        if not is_table_empty():
            print(f"{a+1}. View Data")
            print(f"{a+1}. Update Data")
            print(f"{a+1}. Delete Data")
        print(f"{a+1}. Exit App")
        choice = input("Enter your choice: ")

        match choice:
            case '1':
                add_data()
            case '2':
                view_data()
            case '3':
                update_data()
            case '4':
                delete_data()
            case '5':
                print("Thank you! Have a great day!")
                break
            case _:
                print("Invalid choice!!")

if __name__ == "__main__":
    main()